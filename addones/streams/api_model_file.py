#!/opt/fbi-base/bin/python
# -*- coding: utf-8 -*- 
#xlink的py文件
#filename: api_model_file.xlk
#datetime: 2024-08-30T16:10:58.168928
#copyright: OpenFBI

import sys 
sys.path.append("/opt/openfbi/fbi-bin/driver")
sys.path.append("/opt/openfbi/fbi-bin/lib")
sys.path.append("/opt/openfbi/pylibs")
sys.path.append("../")
import json
from . import *
import threading
try:
	import numpy as np 
	import  pandas as pd
	from avenger.fbiprocesser import *
	from avenger.fglobals import *

except:
	pass



#流和批共享的函数：

#数组到DF
def push_arrays_to_df(arrays,name=""):
	if len(arrays)==0:
		return 0
		
	try:
		#lockP.acquire()
		b  = arrays.copy()
		del arrays[0:len(b)]
		#arrays.clear()
		#lockP.release()

		df = pd.DataFrame(b)
		#设置index 为0
		df['seq19821221'] = 0
		df.set_index('seq19821221',inplace=True)
		if fbi_global.runtime.is_have(name):
			o = fbi_global.runtime.get(name)
			dfs=[o.df,df]
			dfz = pd.concat(dfs,sort=True)
			o.df = dfz
		else:
			o = FbiTable(name,df)
			fbi_global.runtime.put(o)
		if stream["pm_ssdb_printf"]:#用于调试
			o.df.to_feather("/dev/shm/xlink_df:{}:{}.fat".format(stream["name"],name))
	except Exception as e:
		add_error_to_log(stream['name'],'push_arrays_to_df执行出错','DF: %s,原因: %s'%(name,e.__str__()))
#end push_arrays_to_df

#字典到DF
def push_dict_to_df(d,name=""):
	try:
		dd = d.copy() #浅复制，保持不变
		df = pd.DataFrame(data=list(dd.values()),index=list(dd.keys()))
		o = FbiTable(name,df)
		fbi_global.runtime.put(o)
		if stream["pm_ssdb_printf"]:#用于调试
			df.to_feather("/dev/shm/xlink_df:{}:{}.fat".format(stream["name"],name))
	except Exception as e:
		add_error_to_log(stream['name'],'push_dict_to_df执行出错','DF: %s,原因: %s'%(name,e.__str__()))
#end push_dict_to_df


#mysql到DF
def mysql_to_df(a,cols,name=""):
	try:
		df = pd.DataFrame(data=a,columns=cols)
		if stream["pm_ssdb_printf"]:#用于调试
			df.to_feather("/dev/shm/xlink_df:{}:{}.fat".format(stream["name"],name))
	except Exception as e:
		add_error_to_log(stream['name'],'push_mysql_to_df执行出错','DF: %s,原因: %s'%(name,e.__str__()))
	return df
#end mysql_to_df

#参数替换
def replace_ps(p,runtime):	
	for k in runtime.keys:
		p = p.replace(k,runtime.ps[k])	
	return p

#处理单值变量的替换
def deal_sdf(work_space="",prmtv=""):	
	if prmtv.find("$")==-1: #未找到需要替换的变量
		return prmtv

	#仅适用当前工作区
	d = fbi_global.get_runtime().get_cur_ws().workspace	
	
	#处理单值$,不能跨区 add by gjw on 20220507
	keys = list(d.keys())
	#倒序，越长的越在最前面
	keys.sort(key=len,reverse = True)
	
	for k in keys:
		if d[k].type==2:
			try:
				if isinstance(d[k].vue,str):
					prmtv = prmtv.replace("$%s"%(k),d[k].vue)
				else:
					prmtv = prmtv.replace("$%s"%(k),str(d[k].vue))
			except:
				try:
					if isinstance(d[k].vue,str):
						prmtv = prmtv.replace("$%s"%(k),str(d[k].vue,"utf-8"))
				except:
					prmtv = prmtv.replace("$%s"%(k),str(d[k].vue,"gbk"))
	return prmtv

#LastModifyDate:　2024-01-09T10:22:04    Author:   rzc

#LastModifyDate:　2024-01-08T10:07:08    Author:   superFBI

#LastModifyDate:　2024-01-08T09:16:35    Author:   superFBI

#LastModifyDate:　2024-01-06T15:46:26    Author:   superFBI

#LastModifyDate:　2024-01-06T13:54:54    Author:   superFBI

#LastModifyDate:　2024-01-05T09:55:33.565636    Author:   superFBI

#LastModifyDate:　2024-01-05T09:55:10.566403    Author:   superFBI

#LastModifyDate:　2023-12-27T14:10:38.563463    Author:   superFBI

#LastModifyDate:　2023-12-20T11:18:30.880160    Author:   superFBI

#LastModifyDate:　2023-12-13T14:34:09.228842    Author:   superFBI

#LastModifyDate:　2023-12-13T14:11:43.597621    Author:   superFBI

#xlink脚本

#file: api_model_file.xlk

#name: 文件导出模型

#描述： 文件敏感拿数据处理模型

#查看流计算服务

#a = @udf FBI.x_finder3_list

#启动

#a = @udf FBI.x_finder3_start with api_model_file

#停止

#a = @udf FBI.x_finder3_stop with api_model_file

#查询错误日志

#a = load ssdb by ssdb0 query qrange,X_log:api_model_file,0,1000

#查看xlink内部信息, 每５秒更新

#在对象查看器里输入　printf::api_model_file

#断点调试

#debug_on(1)

#初始化

#初始化函数
def Inits():
	try:
		fbi_global.runtime.use_ws("xlink")
	except:
		pass
	stream["meta_name"] = "文件导出模型"
	stream["meta_desc"] = "文件敏感拿数据处理模型"
	#stream["source"]={"link":"127.0.0.1:16379","topic":"model_file","redis":"list"}
	stream["source"]={"unix_udp":"/tmp/model_file"}
	#a=load_ssdb_kv("setting")
	#stream["link"] = a["kfk"]["origin"]["link"]
	#stream["reset"] = a["kfk"]["origin"]["reset"]
	stream["st"]["st_60s"]={"times":60,"fun":"send60"}
	stream["st"]["st_10s"]={"times":10,"fun":"print10"}
	stream["CKH"] = CKH_Client(host="127.0.0.1",port=19000,user="default",password="client")
	#stream["kfk"]={"link":stream["link"],"topic":"api_send","key":""}
	a = load_ssdb_kv("qh_send")["sends"].split(',')
	if "api_model" in a:
		stream["sends4"] = 1
	else:
		stream["sends4"] = 0
	c = load_ssdb_kv("model_config")
	stream["mgcount"] = c["setting"]["model102"]["count"]
	stream["model102_on"] = c["setting"]["switch"]["model102"]
	stream["all_combo"] = c["setting"]["switch"]["all_combo"]
	m102 = c["setting"]["model102"]["whitelist"]
	for i in m102:
		for k in list(i.keys()):
			if not i[k]:
				del i[k]
	stream["model102_conf"] = m102
	srcip_model = load_ssdb_kv("srcip_model_xlk")["data"]
	stream["srcip_model"] = []
	for item in srcip_model:
		stream["srcip_model"].append(item[0])
	s = load_ssdb_kv("monitor_url_xlk")["data"]
	b = []
	for item in s:
		b.append(item[0])
	stream["monitor_url"] = b
	pool["api_model"] = []
	stream["count_excel"] = 0
	stream["count_ok"] = 0
	stream["cmm"] = 0
#end 

#事件处理函数

#事件处理函数
def Events(o,topic=''):
	if o["event_type"] == "fileinfo" and o.get("app_proto","") == "http" and stream["model102_on"]:
		files = "/data/" + o["fileinfo"]["file_path"]
		t_path = "/dev/shm/znsm/"
		if not os.path.exists(t_path):
			os.mkdir(t_path)
		#判断文件大小
		max_size=4*1024*1024*1024
		start = o.get("http",{}).get("start","")
		end = o.get("http",{}).get("end","")
		if os.path.exists(files):
			file_size=getDocSize(files)
			#判断size是否为空
			if file_size is not None and "1970-01-01" not in start and "1970-01-01" not in end:
				if file_size < max_size:
					filename=o.get("fileinfo",{}).get("filename")
					if ".xlsx" in filename or ".xls" in filename:
						#读取excel的行号
						try:
							row_counts,sheet_counts=read_row(files,filename)
						except:
							row_counts = 0
							sheet_counts = 0
						if row_counts and sheet_counts:
							stream["count_excel"] += 1
							#printf("row",row_counts)
							#printf("sheet",sheet_counts)
							try:
								filename = filename.encode('latin1').decode('gb2312')
							except:
								filename = filename
							filename = parse.unquote(filename)
							#printf("filename",filename)
							count1 = row_counts - sheet_counts
							if count1 > stream["mgcount"]:
								stream["count_ok"] += 1
								srcip = o.get("src_ip")
								srcall = 1
								if srcip in stream["srcip_model"] or srcall:
									http = o.get("http")
									uri = http.get("url","")
									hostname = http.get("hostname")
									if hostname:
										app = hostname
										url = "http://" + hostname + uri.split("?")[0]
									else:
										app = srcip
										url = "http://" + srcip + uri.split("?")[0]
									if stream["all_combo"] or url in stream["monitor_url"] or srcall:
										http_api = {}
										http_api["url"] = url
										http_api["srcip"] = o.get("dest_ip")
										http_api["srcport"] = o.get("src_port")
										http_api["dstip"] = srcip
										http_api["dstport"] = o.get("dest_port")
										ps = 0
										for i in stream["model102_conf"]:
											a = 0
											for k in i:
												if http_api[k] == i[k]:
													a = a + 1
											if len(i) == a:
												ps = 1
												break
										if not stream["model102_conf"] or ps:
											http_api["timestamp"] = iso_to_datetime(o.get("timestamp"))
											http_api["url_a"] = url
											http_api["account"] = ""
											http_api["real_ip"] = ""
											http_api["app"] = app
											http_api["id"] = xlink_uuid(1)
											http_api["type"] = 102
											http_api["level"] = 2
											http_api["proof"] = str(o.get("flow_id"))
											http_api["desc"] = "终端通过接口进行文件导出"
											http_api["message"] = "终端“" + o.get("dest_ip") + "”在接口" + url + "中导出文件含有" + str(count1) + "条数据"
											proofs = {}
											proofs["判定标准"] = "文件导出行为:终端通过接口进行文件导出，excel文件中数据条数超过阈值限制进行告警"
											proofs["接口"] = url
											proofs["终端"] = o.get("dest_ip")
											proofs["文件名"] = filename
											proofs["MD5"] = o.get("fileinfo",{}).get("md5","")
											proofs["SHA256"] = o.get("fileinfo",{}).get("sha256","")
											proofs["单次获取阈值"] = stream["mgcount"]
											proofs["本次获取数量"] = count1
											proofs["结果"] = "终端通过接口进行文件导出"
											proofs["证据"] = {}
											proofs["证据"]["时间"] = "HTTP协议ID"
											proofs["证据"][str(o.get("timestamp"))] = o.get("flow_id")
											row_data9 = []
											row_data10 = []
											if filename.endswith(".xlsx"):
												# openpyxl
												file2 = files + ".xlsx"
												if not os.path.exists(file2):
													os.rename(files, file2)
												workbook = load_workbook(file2)
												#printf("read_file",workbook)
												worksheet = ""
												try:
													for i in range(10):
														worksheet = workbook[workbook.sheetnames[i]]
														i = worksheet.max_row
														real_max_row = 0
														while i > 0:
															row_dict = {i.value for i in worksheet[i]}
															if row_dict == {None}:
																i = i - 1
															else:
																real_max_row = i
																break
														if real_max_row > 10:
															break
												except:
													pass
												for cell in worksheet[9]:
													row_data9.append(cell.value)
												for cell in worksheet[9]:
													row_data10.append(cell.value)
												workbook.close()
												try:
													os.rename(file2, files)
												except:
													stream["cmm"] += 1
											if filename.endswith(".xls"):
												# xlrd
												file2 = files + ".xls"
												if not os.path.exists(file2):
													os.rename(files, file2)
												workbook = open_workbook(file2)
												#printf("read_file",workbook)
												worksheet = ""
												max_row = ""
												try:
													for i in range(10):
														worksheet = workbook.sheets()[i]
														max_row = worksheet.nrows
														if max_row > 10:
															break
												except:
													pass
												row_data9 = worksheet.row_values(9)
												row_data10 = worksheet.row_values(10)
												workbook.release_resources()
												try:
													os.rename(file2, files)
												except:
													stream["cmm"] += 1
											proofs["证据"]["示例1"] = row_data9
											proofs["证据"]["示例2"] = row_data10
											http_api["proofs"] = proofs
											to_pool("api_model",http_api)
											if stream["sends4"]:
												ssss = deepcopy(http_api)
												ssss["event_type"] = "model"
												#to_kfk(ssss)
												to_json_file("/data/syslog_file/eve",ssss)
#end 

#系统定时函数

#系统定时函数，st为时间戳 
def print10(st):
	store_ckh(pool["api_model"],"api_model")
	printf("count","匹配到excel：%d===符合:%d===重命名失败:%d"%(stream["count_excel"],stream["count_ok"],stream["cmm"]))
	
#end 

#系统定时函数，st为时间戳 
def send60(st):
	a = load_ssdb_kv("qh_send")["sends"].split(',')
	if "api_model" in a:
		stream["sends4"] = 1
	else:
		stream["sends4"] = 0
	c = load_ssdb_kv("model_config")
	stream["mgcount"] = c["setting"]["model102"]["count"]
	stream["all_combo"] = c["setting"]["switch"]["all_combo"]
	stream["model102_on"] = c["setting"]["switch"]["model102"]
	m102 = c["setting"]["model102"]["whitelist"]
	for i in m102:
		for k in list(i.keys()):
			if not i[k]:
				del i[k]
	stream["model102_conf"] = m102
	srcip_model = load_ssdb_kv("srcip_model_xlk")["data"]
	src = []
	for item in srcip_model:
		src.append(item[0])
	stream["srcip_model"] = src
	s = load_ssdb_kv("monitor_url_xlk")["data"]
	b = []
	for item in s:
		b.append(item[0])
	stream["monitor_url"] = b
#end 

#自定义python函数，必须有参数,可以在初始化函数，事件处理函数，和定时函数中使用
def xlink_uuid(x):
	return str(time.time_ns())
#end 

#需要额外引入的包

#需要引入的包 
import sys
import gc
import base64
from copy import deepcopy
from urllib import parse
from un_file import *
from openpyxl import load_workbook
from xlrd import open_workbook
#end 

#udf

#end 
